#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Crypto Move Readiness Scanner
Author: ChatGPT
Description:
  - Fetch top-N crypto assets by market cap (CoinGecko).
  - Pull OHLCV from major exchanges via CCXT (Binance preferred, fallback OKX/Coinbase).
  - Compute indicators (RSI, MFI, Bollinger bandwidth, ATR%, Volume Z-score).
  - Try to pull current funding rates from Binance USDT-margined futures for available tickers.
  - Build a composite "Move Readiness Score" highlighting assets likely to move strongly soon.
  - Save an Excel report with Summary + per-asset sheets.

Usage:
  python main.py --tf 1h --days 90 --top 50
"""
import argparse
import time
import math
import sys
import traceback
from typing import List, Dict, Any, Optional, Tuple

import requests
import pandas as pd
import numpy as np

# Indicators via pandas_ta (pure python)
try:
    import pandas_ta as ta
except Exception as e:
    print("pandas_ta is required. Please ensure 'pandas-ta' is installed.")
    raise

# CCXT for exchange data
import ccxt

# Excel formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

COINGECKO_API = "https://api.coingecko.com/api/v3"

DEFAULT_EXCHANGES = ["binance", "okx", "coinbase"]
FUTURES_EXCHANGE_ID = "binanceusdm"  # Binance USDT-M Perp

def get_top_coins(n=50) -> pd.DataFrame:
    # Use CoinGecko /coins/markets to get top by market cap
    url = f"{COINGECKO_API}/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": n,
        "page": 1,
        "price_change_percentage": "24h",
        "locale": "en"
    }
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    data = r.json()
    df = pd.json_normalize(data)
    # Keep necessary columns
    keep = ["id","symbol","name","market_cap_rank","market_cap","current_price","total_volume"]
    return df[keep].rename(columns={"symbol":"base_symbol"})

def boot_exchange(eid: str):
    cls = getattr(ccxt, eid)
    ex = cls({"enableRateLimit": True, "options": {"defaultType": "spot"}})
    ex.load_markets()
    return ex

def boot_futures_exchange(eid: str):
    cls = getattr(ccxt, eid)
    ex = cls({"enableRateLimit": True, "options": {"defaultType": "future"}})
    ex.load_markets()
    return ex

def pick_symbol_on_exchange(ex, base: str, quote: str = "USDT") -> Optional[str]:
    # Prefer base/USDT, fallback base/USD
    candidates = [f"{base}/{quote}", f"{base}/USD"]
    for s in candidates:
        if s in ex.markets:
            return s
    # try uppercase/lowercase mismatch
    for m in ex.markets:
        if m.split("/")[0].upper() == base.upper() and m.split("/")[1] in [quote, "USD"]:
            return m
    return None

def fetch_ohlcv_any(exchanges: List[str], base: str, tf: str, since_ms: int, limit: int) -> Optional[pd.DataFrame]:
    last_err = None
    for eid in exchanges:
        try:
            ex = boot_exchange(eid)
            sym = pick_symbol_on_exchange(ex, base)
            if not sym:
                continue
            ohlcv = ex.fetch_ohlcv(sym, timeframe=tf, since=since_ms, limit=limit)
            if not ohlcv:
                continue
            df = pd.DataFrame(ohlcv, columns=["timestamp","open","high","low","close","volume"])
            df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True).dt.tz_convert("UTC")
            df["exchange"] = eid
            df["symbol"] = sym
            return df
        except Exception as e:
            last_err = e
            time.sleep(0.2)
            continue
    if last_err:
        print(f"[warn] OHLCV fetch failed for {base}: {last_err}")
    return None

def compute_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.set_index("timestamp", inplace=True)
    # Basic indicators
    df["rsi14"] = ta.rsi(df["close"], length=14)
    # Money Flow Index requires high, low, close, volume
    df["mfi14"] = ta.mfi(high=df["high"], low=df["low"], close=df["close"], volume=df["volume"], length=14)
    bb = ta.bbands(df["close"], length=20)
    if bb is not None and "BBL_20_2.0" in bb and "BBU_20_2.0" in bb:
        df["bb_width"] = (bb["BBU_20_2.0"] - bb["BBL_20_2.0"]) / df["close"]
    else:
        df["bb_width"] = np.nan

    # ATR% over close
    atr = ta.atr(high=df["high"], low=df["low"], close=df["close"], length=14)
    df["atr_pct"] = atr / df["close"]

    # Volume Z-score (rolling 20)
    df["vol_z"] = (df["volume"] - df["volume"].rolling(20).mean()) / df["volume"].rolling(20).std(ddof=0)

    # Momentum: recent % change (close vs close n-bars ago)
    df["chg_24"] = df["close"].pct_change(24)  # approx 1d on 1h tf

    df.reset_index(inplace=True)
    return df

def get_current_funding_rate(binance_fut, base: str) -> Optional[float]:
    # Try common perpetual symbols: BASE/USDT:USDT for ccxt unified
    # Fallback to raw markets lookup
    try_syms = []
    # First try exact unified perp
    for m, spec in binance_fut.markets.items():
        if spec.get("type") == "future" and spec.get("linear"):
            # spec['base'] equals base?
            if str(spec.get("base","")).upper() == base.upper() and "PERPETUAL" in str(spec.get("info",{})).upper():
                try_syms.append(m)
    try_syms = list(dict.fromkeys(try_syms))  # unique preserve order

    for sym in try_syms:
        try:
            fr = binance_fut.fetch_funding_rate(sym)
            if isinstance(fr, dict) and "fundingRate" in fr:
                return float(fr["fundingRate"])
            if isinstance(fr, dict) and "info" in fr and "lastFundingRate" in fr["info"]:
                return float(fr["info"]["lastFundingRate"])
        except Exception:
            continue
    return None

def normalize_score(val, lo, hi):
    if pd.isna(val):
        return 0.0
    return max(-1.0, min(1.0, 2 * ( (val - lo) / (hi - lo + 1e-12) ) - 1 ))

def build_composite_row(base: str, indf: pd.DataFrame, funding: Optional[float]) -> Dict[str, Any]:
    # Pick last complete row
    last = indf.dropna().tail(1)
    if last.empty:
        last = indf.tail(1)
    row = last.iloc[0].to_dict()

    # Normalize components
    rsi_score = 0.0
    if not pd.isna(row.get("rsi14")):
        # Overbought/oversold tension: distance from 50
        rsi_score = (row["rsi14"] - 50.0) / 50.0  # -1..+1

    mfi_score = 0.0
    if not pd.isna(row.get("mfi14")):
        mfi_score = (row["mfi14"] - 50.0) / 50.0

    bb_score = normalize_score(row.get("bb_width"), lo=0.005, hi=0.10)  # wider bands suggest move potential

    vol_score = 0.0
    if not pd.isna(row.get("vol_z")):
        # Clamp into [-2, +2], then scale to [-1, +1]
        v = max(-2.0, min(2.0, row["vol_z"]))
        vol_score = v / 2.0

    atr_score = normalize_score(row.get("atr_pct"), lo=0.005, hi=0.10)

    fund_score = 0.0
    if funding is not None:
        # Extreme funding (pos or neg) indicates skew/risk of squeeze
        # Normalize -0.05 .. +0.05 to -1..+1
        funding = max(-0.05, min(0.05, funding))
        fund_score = funding / 0.05

    # Weighting (can tune via config)
    weights = {
        "rsi": 0.25,
        "mfi": 0.15,
        "bb": 0.20,
        "vol": 0.20,
        "atr": 0.10,
        "fund": 0.10
    }
    composite = (weights["rsi"] * rsi_score +
                 weights["mfi"] * mfi_score +
                 weights["bb"]  * bb_score +
                 weights["vol"] * vol_score +
                 weights["atr"] * atr_score +
                 weights["fund"]* fund_score)

    out = {
        "asset": base.upper(),
        "exchange": row.get("exchange",""),
        "symbol": row.get("symbol",""),
        "timestamp": row.get("timestamp"),
        "close": row.get("close"),
        "volume": row.get("volume"),
        "rsi14": row.get("rsi14"),
        "mfi14": row.get("mfi14"),
        "bb_width": row.get("bb_width"),
        "atr_pct": row.get("atr_pct"),
        "vol_z": row.get("vol_z"),
        "chg_24": row.get("chg_24"),
        "funding_rate": funding,
        "move_readiness_score": composite
    }
    return out

def write_excel(summary_df: pd.DataFrame, per_asset: Dict[str, pd.DataFrame], out_path: str):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Summary
        summary_df.sort_values("move_readiness_score", ascending=False).to_excel(writer, sheet_name="Summary", index=False)
        # Each asset sheet (optional: only top 15 to keep file smaller)
        for asset, df in per_asset.items():
            df.to_excel(writer, sheet_name=asset[:31], index=False)

    # Styling Summary sheet
    wb = load_workbook(out_path)
    ws = wb["Summary"]
    # Header styling
    header_fill = PatternFill(start_color="1f4e78", end_color="1f4e78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
            except:
                val = ""
            if val:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(40, max(12, max_len + 2))
    wb.save(out_path)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--tf", default="1h", help="Timeframe (e.g., 1h, 4h, 1d)")
    parser.add_argument("--days", type=int, default=90, help="History window in days")
    parser.add_argument("--top", type=int, default=50, help="Top-N by market cap")
    parser.add_argument("--outfile", default="crypto_move_readiness_report.xlsx", help="Excel output filename")
    args = parser.parse_args()

    print(f"Params: tf={args.tf}, days={args.days}, top={args.top}")

    # Fetch top coins
    print("Fetching top coins from CoinGecko...")
    top_df = get_top_coins(args.top)

    # Prepare time range
    tf_map = {"1h": 60, "4h": 240, "1d": 1440, "1m": 1, "5m": 5, "15m": 15}
    if args.tf not in tf_map:
        print(f"Unsupported timeframe {args.tf}. Try one of: {list(tf_map.keys())}")
        sys.exit(1)
    minutes = tf_map[args.tf]
    bars = int((args.days * 24 * 60) / minutes)
    # Extra bars for indicators warmup
    bars = min(2000, bars + 50)
    since_ms = int((pd.Timestamp.utcnow() - pd.Timedelta(minutes=minutes * bars)).timestamp() * 1000)

    # Boot futures exchange once
    try:
        binance_fut = boot_futures_exchange(FUTURES_EXCHANGE_ID)
    except Exception as e:
        print(f"[warn] Could not boot futures exchange {FUTURES_EXCHANGE_ID}: {e}")
        binance_fut = None

    results = []
    per_asset_frames = {}

    for _, row in top_df.iterrows():
        base = str(row["base_symbol"]).upper()
        # Common stablecoin noise skip
        if base in ["USDT","USDC","FDUSD","DAI","TUSD"]:
            continue
        print(f"Processing {base} ...")
        try:
            ohlcv_df = fetch_ohlcv_any(DEFAULT_EXCHANGES, base, args.tf, since_ms, limit=bars)
            if ohlcv_df is None or len(ohlcv_df) < 50:
                print(f"[skip] No OHLCV for {base}")
                continue
            indf = compute_indicators(ohlcv_df)

            # Funding
            fr = None
            if binance_fut is not None:
                try:
                    fr = get_current_funding_rate(binance_fut, base)
                except Exception as e:
                    fr = None

            comp_row = build_composite_row(base, indf, fr)
            results.append(comp_row)

            # Keep compact per-asset frame (last 300 rows)
            per_asset_frames[base] = indf.tail(300)

        except Exception as e:
            print(f"[warn] {base} failed: {e}")
            traceback.print_exc()
            continue

    if not results:
        print("No results to write.")
        return

    summary_df = pd.DataFrame(results)
    # Also include absolute score for ranking (magnitude)
    summary_df["abs_score"] = summary_df["move_readiness_score"].abs()

    # Save Excel
    out_path = args.outfile
    write_excel(summary_df, per_asset_frames, out_path)
    print(f"Saved report -> {out_path}")

if __name__ == "__main__":
    main()
